"""
RAG
this file creates the long term agent, it is done by instrument callee RAG
create Index
It reads the Word script files, splits them into chunks, generates embeddings, and stores them in Chroma.

Folder structure:

scripts/
    Each .docx file = one video

metadata.csv    (optional but recommended)
    file,title,url

If metadata.csv is not present, the filename is used as the video title.

Important: This step takes several minutes. Don't run it while recording.

Run it before recording, and in the video, only show the code.
    python ingest.py
"""

import csv
import os
import time
from pathlib import Path

from docx import Document as Docx
from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_chroma import Chroma

from config import embedder, CHROMA_DIR, COLLECTION

SCRIPTS_DIR = Path("scripts")
META_CSV = Path("metadata.csv")
META_XLSX = Path("metadata.xlsx")
# The spoken script has natural paragraph breaks, but no headings.

# 900 characters is roughly one minute of speech — a good semantic unit.
CHUNK_SIZE = 900
CHUNK_OVERLAP = 150

SEPARATORS = ["\n\n", "\n", "؟ ", ". ", "، ", " ", ""]


# # Excel on Windows with Persian settings usually saves files using cp1256, not UTF-8.ا
ENCODINGS = ["utf-8-sig", "utf-8", "cp1256", "cp1252"]


def _read_text(path):
    """read the file with any coding  ."""
    for enc in ENCODINGS:
        try:
            text = path.read_text(encoding=enc)
            if enc not in ("utf-8-sig", "utf-8"):
                print(f"  [!] metadata.csv  with coding csv  {enc} is read  .")
            return text
        except (UnicodeDecodeError, LookupError):
            continue
    raise SystemExit("metadata.csv with any coding was read ا    .")


def _rows_from_xlsx():
    """ read Excel directly, it doesnt have any coding issueا  .    ."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("    to read from excel instal this   :  pip install openpyxl")

    wb = load_workbook(META_XLSX, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(r):
            continue
        out.append({header[i]: (str(r[i]).strip() if i < len(r) and r[i] else "")
                    for i in range(len(header))})
    print(f"metadata.xlsx  is read  — {len(out)} row ")
    return out


def _rows_from_csv():
    import io
    text = _read_text(META_CSV)
    first = text.splitlines()[0] if text.splitlines() else ""
    delim = ";" if first.count(";") > first.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    out = [{(k or "").strip().lower(): (v or "").strip() for k, v in row.items()}
           for row in reader]
    print(f"metadata.csv  is read  — {len(out)} row ")
    return out


def read_metadata():
    """ file name   → (  video title or link ).    first excel then csv ."""
    if META_XLSX.exists():
        rows = _rows_from_xlsx()
    elif META_CSV.exists():
        rows = _rows_from_csv()
    else:
        print("if the file was not metadata ل دthe name of file is used as  title ")
        return {}

    if rows and "file" not in rows[0]:
        raise SystemExit(
            "was not  found the file column \n"
            f" excisted column  : {list(rows[0].keys())}\n"
            "needs 3 column with these names       : file  title  url"
        )

    meta = {}
    for r in rows:
        name = r.get("file", "")
        if name:
            meta[name] = (r.get("title", ""), r.get("url", ""))
    return meta


def read_docx(path):
    doc = Docx(path)
    parts = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(parts)


def load_all():
    if not SCRIPTS_DIR.exists():
        raise SystemExit("folder  scripts  if its not fould create a file and insert the new files in it  .")
    meta = read_metadata()
    files = sorted(SCRIPTS_DIR.glob("*.docx"))
    files = [f for f in files if not f.name.startswith("~$")]  # temporaary files created for enetering 

    if not files:
        raise SystemExit("there is no  .docx  in folder   scripts is not .")

    docs = []
    for f in files:
        text = read_docx(f)
        if len(text) < 200:
            print(f"  [!] {f.name}     almost empty its failed .")
            continue

        title, url = meta.get(f.name, (f.stem, ""))
        if not title:
            title = f.stem

        docs.append(Document(
            page_content=text,
            metadata={"title": title, "url": url, "file": f.name},
        ))
        print(f"  {f.name}  →  {len(text.split()):,} word   |  {title}")

    return docs


def main():
    print("=" * 60)
    print("    create index script Moji ")
    print("=" * 60)

    docs = load_all()
    print(f"\n{len(docs)}   video is read .")

    splitter = RecursiveCharacterTextSplitter(
        chunk_size=CHUNK_SIZE,
        chunk_overlap=CHUNK_OVERLAP,
        separators=SEPARATORS,
        length_function=len,
    )
    chunks = splitter.split_documents(docs)

    # We keep the chunk number so it can later be mapped to an approximate timestamp.
    per_video = {}
    for c in chunks:
        t = c.metadata["title"]
        per_video[t] = per_video.get(t, 0) + 1
        c.metadata["chunk"] = per_video[t]

    print(f"{len(chunks):,}   chunck is created .")
    print(f"average  {len(chunks) / len(docs):.0f} chunck for each video .\n")

    print("Embeding model is loading     ...")
    emb = embedder()

    if os.path.exists(CHROMA_DIR):
        print("previous index is  cleared     ...")
        import shutil
        shutil.rmtree(CHROMA_DIR)

    print("Embedding ..this process takes sevral minuets   .   ...\n")
    t = time.time()

    store = Chroma(
        collection_name=COLLECTION,
        embedding_function=emb,
        persist_directory=CHROMA_DIR,
    )

    BATCH = 64
    for i in range(0, len(chunks), BATCH):
        store.add_documents(chunks[i:i + BATCH])
        done = min(i + BATCH, len(chunks))
        print(f"  {done:,} / {len(chunks):,}", end="\r", flush=True)

    dt = time.time() - t
    print(f"\n\ finished  in     {dt / 60:.1f} minutes.")
    print(f" index in  {CHROMA_DIR}  is saved .")

    #   a fast test 
    print("\n" + "-" * 60)
    print("restoring test  :")
    hits = store.similarity_search(" what is Embedding", k=3)
    for h in hits:
        print(f"  [{h.metadata['title']} — chunck {h.metadata['chunk']}]")
        print(f"    {h.page_content[:90]} ...")


if __name__ == "__main__":
    main()